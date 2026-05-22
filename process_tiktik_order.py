"""
TikTok Shop → FlashShip XLSX Import Generator
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import csv
import json
import os
import sys
from datetime import date
from urllib.parse import urlparse
import openpyxl

# ─── Config ────────────────────────────────────────────────────────────────────
# When frozen by PyInstaller, bundled data lives in sys._MEIPASS.
# The output file is always written next to the running executable / .app.
if getattr(sys, "frozen", False):
    _BUNDLE_DIR = sys._MEIPASS                          # bundled resources
else:
    _BUNDLE_DIR = os.path.dirname(os.path.abspath(__file__))

# Always save exported files to the Desktop
SCRIPT_DIR   = os.path.join(os.path.expanduser("~"), "Desktop")

MAPPING_FILE = os.path.join(_BUNDLE_DIR, "flashship_mapping.json")

CANCELLED_STATUSES = {"Cancelled", "Seller Cancel", "Cancel Request", "Unpaid"}
VOUCHER_PREFIX = "Spend $"
SHIP_STATUS    = "To ship"
PLACEHOLDER    = "✏  Nhấp để nhập link..."

FLASHSHIP_COLUMNS = [
    "Order ID", "Shipping method", "Customer's name", "Email", "Phone",
    "Country", "State", "Address line 1", "Address line 2", "City", "Zip",
    "Link Label", "Quantity", "Variant ID",
    "Design front", "Design back",
    "Design Left Hand", "Design Right Hand", "Design Neck", "Design Hood",
    "Design Pocket", "Design Neck Label Inner", "Special Print",
    "Front Extra Large", "Back Extra Large", "Left Extra Large", "Right Extra Large",
    "Mockup Front", "Mockup Back", "Mockup Left Hand", "Mockup Right Hand",
    "Mockup Neck", "Mockup Hood", "Mockup Pocket", "Mockup Neck Label Inner",
    "Product Note", "DTF/DTG", "Card Code",
]

STATE_MAP = {
    "Alabama":"AL","Alaska":"AK","Arizona":"AZ","Arkansas":"AR","California":"CA",
    "Colorado":"CO","Connecticut":"CT","Delaware":"DE","Florida":"FL","Georgia":"GA",
    "Hawaii":"HI","Idaho":"ID","Illinois":"IL","Indiana":"IN","Iowa":"IA",
    "Kansas":"KS","Kentucky":"KY","Louisiana":"LA","Maine":"ME","Maryland":"MD",
    "Massachusetts":"MA","Michigan":"MI","Minnesota":"MN","Mississippi":"MS",
    "Missouri":"MO","Montana":"MT","Nebraska":"NE","Nevada":"NV","New Hampshire":"NH",
    "New Jersey":"NJ","New Mexico":"NM","New York":"NY","North Carolina":"NC",
    "North Dakota":"ND","Ohio":"OH","Oklahoma":"OK","Oregon":"OR","Pennsylvania":"PA",
    "Rhode Island":"RI","South Carolina":"SC","South Dakota":"SD","Tennessee":"TN",
    "Texas":"TX","Utah":"UT","Vermont":"VT","Virginia":"VA","Washington":"WA",
    "West Virginia":"WV","Wisconsin":"WI","Wyoming":"WY","District of Columbia":"DC",
}

# ─── Helpers ────────────────────────────────────────────────────────────────────
def load_mapping():
    if not os.path.exists(MAPPING_FILE):
        return {}, {}, {}
    with open(MAPPING_FILE, encoding="utf-8") as f:
        data = json.load(f)
    return data.get("variant_map", {}), data.get("size_fix", {}), data.get("color_fix", {})

def lookup_variant(variation, variant_map, size_fix, color_fix):
    parts = [p.strip() for p in variation.split(",")]
    if len(parts) == 2:
        color, size = parts[0], parts[1]
        color = color_fix.get(color, color)
        size  = size_fix.get(size, size)
        fixed = f"{color}, {size}"
    else:
        fixed = variation
    return variant_map.get(fixed), fixed

def is_valid_url(value):
    if not value:
        return True
    try:
        r = urlparse(value)
        return r.scheme in ("http", "https") and bool(r.netloc)
    except Exception:
        return False

# ─── Filter rows ────────────────────────────────────────────────────────────────
def filter_rows(rows, variant_map, size_fix, color_fix):
    items = []
    for row in rows:
        status    = row.get("Order Status", "").strip()
        variation = row.get("Variation", "").strip()
        order_id  = row.get("Order ID", "").strip()
        if not order_id:
            continue
        if status in CANCELLED_STATUSES:
            continue
        if variation.startswith(VOUCHER_PREFIX):
            continue
        if status != SHIP_STATUS:
            continue

        vid, fixed_var = lookup_variant(variation, variant_map, size_fix, color_fix)

        raw_phone = row.get("Phone#", row.get("Phone", "")).strip()
        digits = "".join(c for c in raw_phone if c.isdigit())
        if digits.startswith("1") and len(digits) == 11:
            digits = digits[1:]

        state_raw = row.get("Province", row.get("State", "")).strip()
        state     = STATE_MAP.get(state_raw, state_raw)

        items.append({
            "order_id":        order_id,
            "order_date":      row.get("Created Time", row.get("Order Creation Date", "")).strip(),
            "customer":        row.get("Recipient", row.get("Ship to name", "")).strip(),
            "variation":       variation,
            "fixed_variation": fixed_var,
            "variant_id":      vid,
            "quantity":        row.get("Quantity", "1").strip(),
            "status_note":     "" if vid else "❌ Không có Variant ID. Phải làm order thủ công.",
            "partial_lock":    False,
            "phone":           digits,
            "state":           state,
            "address1":        row.get("Ship to address",  row.get("Address line 1", "")).strip(),
            "address2":        row.get("Ship to address2", row.get("Address line 2", "")).strip(),
            "city":            row.get("City", "").strip(),
            "zip":             row.get("Zipcode", row.get("Zip", "")).strip(),
            "link_label":      "",
            "design_front":    "",
            "design_back":     "",
            "mockup_front":    "",
            "mockup_back":     "",
        })

    mark_partial_orders(items)
    return items


def mark_partial_orders(items):
    """
    For any order_id where at least one item has no variant_id,
    mark ALL items in that order as partial_lock=True so the entire
    order is blocked from export — no partial orders allowed.
    """
    from collections import defaultdict
    order_has_missing = defaultdict(bool)
    for item in items:
        if not item["variant_id"]:
            order_has_missing[item["order_id"]] = True

    for item in items:
        if order_has_missing[item["order_id"]] and item["variant_id"]:
            # This item has a variant_id but a sibling in the same order doesn't
            item["partial_lock"] = True
            item["status_note"]  = (
                "🟠 Đơn hàng có sản phẩm khác chưa có Variant ID. "
                "Không thể xuất một phần đơn hàng."
            )


def get_partial_export_violations(items, checked_indices):
    """
    Return a list of violation strings for orders that have only *some* of their
    exportable items checked.  An empty list means the selection is safe to export.

    Parameters
    ----------
    items           : list of item dicts (from filter_rows / mark_partial_orders)
    checked_indices : iterable of integer indices that the user has ticked

    An "exportable" item is one where variant_id is truthy AND partial_lock is False.
    """
    from collections import defaultdict
    checked_set = set(checked_indices)

    order_exportable = defaultdict(list)   # order_id -> [exportable indices]
    order_checked    = defaultdict(list)   # order_id -> [checked exportable indices]

    for i, item in enumerate(items):
        if not item["variant_id"] or item.get("partial_lock"):
            continue  # already fully blocked
        order_exportable[item["order_id"]].append(i)
        if i in checked_set:
            order_checked[item["order_id"]].append(i)

    violations = []
    for oid, exportable in order_exportable.items():
        checked = order_checked.get(oid, [])
        if 0 < len(checked) < len(exportable):
            sample = items[exportable[0]].get("customer") or oid
            violations.append(
                f"  • {sample} ({oid}) — đã chọn {len(checked)}/{len(exportable)} sản phẩm"
            )
    return violations


def build_flashship_row(item):
    """
    Build the FlashShip import row dict for a single exportable item.
    The caller is responsible for ensuring item["variant_id"] is set.
    """
    row = {col: "" for col in FLASHSHIP_COLUMNS}
    row["Order ID"]        = "HD - " + item["order_id"]
    row["Shipping method"] = "1"
    row["Customer's name"] = item["customer"]
    row["Phone"]           = item["phone"]
    row["Country"]         = "US"
    row["State"]           = item["state"]
    row["Address line 1"]  = item["address1"]
    row["Address line 2"]  = item["address2"]
    row["City"]            = item["city"]
    row["Zip"]             = item["zip"]
    row["Link Label"]      = item["link_label"]
    row["Quantity"]        = item["quantity"]
    row["Variant ID"]      = item["variant_id"]
    row["Design front"]    = item["design_front"]
    row["Design back"]     = item["design_back"]
    row["Mockup Front"]    = item["mockup_front"]
    row["Mockup Back"]     = item["mockup_back"]
    row["DTF/DTG"]         = "3"
    return row


def parse_order_date(date_str):
    """
    Parse a TikTok order date string (various formats) into a ``datetime.date``.
    Returns ``None`` if the string cannot be parsed.
    """
    from datetime import datetime
    date_str = date_str.strip()
    for fmt in (
        "%m/%d/%Y %I:%M:%S %p",   # TikTok: 05/20/2026 7:43:26 PM
        "%m/%d/%Y %H:%M:%S",      # 24-hour fallback
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
    ):
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
    return None


# ─── App ────────────────────────────────────────────────────────────────────────
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("TikTok Shop → FlashShip CSV Generator")
        self.geometry("1400x700")
        self.minsize(900, 500)
        self.resizable(True, True)

        self.variant_map, self.size_fix, self.color_fix = load_mapping()
        self.items       = []
        self.check_vars  = []
        self._edit_entry     = None
        self._edit_iid       = None
        self._edit_col_index = None
        self._edit_col_name  = None

        self._build_ui()

    # ── UI ──────────────────────────────────────────────────────────────────────
    def _build_ui(self):
        toolbar = tk.Frame(self, pady=6, padx=10)
        toolbar.pack(fill="x")

        tk.Button(toolbar, text="📂  Mở file TikTok CSV", command=self._open_csv,
                  font=("Helvetica", 13, "bold"), bg="#e8e8e8", fg="black",
                  relief="flat", padx=12, pady=6).pack(side="left")

        self.export_btn = tk.Button(toolbar, text="💾  Xuất file FlashShip XLSX",
                                    command=self._export,
                                    font=("Helvetica", 13, "bold"),
                                    bg="#d0e8d0", fg="black",
                                    relief="flat", padx=12, pady=6,
                                    state="disabled")
        self.export_btn.pack(side="left", padx=10)

        tk.Button(toolbar, text="☑ Tất cả", command=self._select_all,
                  relief="flat", padx=8).pack(side="left", padx=(20, 2))
        tk.Button(toolbar, text="☐ Bỏ chọn", command=self._select_none,
                  relief="flat", padx=8).pack(side="left")

        self.status_var = tk.StringVar(value="Mở file TikTok CSV để bắt đầu.")
        tk.Label(toolbar, textvariable=self.status_var,
                 font=("Helvetica", 11), fg="#555").pack(side="right")

        legend = tk.Frame(self, padx=10)
        legend.pack(fill="x")
        tk.Label(legend,
                 text="💡 Nhấp vào ô Link Label / Mặt trước / Mặt sau để chỉnh sửa.   "
                      "� Hàng đỏ = thiếu Variant ID.   "
                      "🟠 Hàng cam = đơn hàng chưa đầy đủ.   (Cả hai đều bị khóa.)",
                 font=("Helvetica", 10), fg="#555").pack(side="left")

        # ── Table ───────────────────────────────────────────────────────────────
        frame = tk.Frame(self)
        frame.pack(fill="both", expand=True, padx=10, pady=(4, 10))

        cols = ("check", "order_date", "order_id", "customer",
                "variation", "variant_id", "qty",
                "link_label", "design_front", "design_back",
                "mockup_front", "mockup_back", "note")
        self.tree = ttk.Treeview(frame, columns=cols, show="headings",
                                 selectmode="browse")

        col_cfg = [
            ("check",        "✔",                   58,  "center"),
            ("order_date",   "Ngày đặt",            120,  "w"),
            ("order_id",     "Mã đơn hàng",         170,  "w"),
            ("customer",     "Khách hàng",           140,  "w"),
            ("variation",    "Sản phẩm",             180,  "w"),
            ("variant_id",   "Variant ID",            90,  "center"),
            ("qty",          "SL",                    40,  "center"),
            ("link_label",   "Link Label  ✏",        210,  "w"),
            ("design_front", "Mặt trước  ✏",         210,  "w"),
            ("design_back",  "Mặt sau  ✏",           210,  "w"),
            ("mockup_front", "Mockup Mặt trước  ✏",  210,  "w"),
            ("mockup_back",  "Mockup Mặt sau  ✏",    210,  "w"),
            ("note",         "Ghi chú",               320,  "w"),
        ]
        for cid, heading, width, anchor in col_cfg:
            self.tree.heading(cid, text=heading)
            self.tree.column(cid, width=width, minwidth=30, anchor=anchor,
                             stretch=cid in ("link_label","design_front","design_back",
                                             "mockup_front","mockup_back",
                                             "variation","note"))

        vsb = ttk.Scrollbar(frame, orient="vertical",   command=self.tree.yview)
        hsb = ttk.Scrollbar(frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)

        self.tree.tag_configure("locked",         background="#ffe0e0", foreground="#888888")
        self.tree.tag_configure("partial_locked", background="#fff0d0", foreground="#888888")
        self.tree.tag_configure("url_err",        background="#ffd0d0", foreground="#000000")
        self.tree.tag_configure("placeholder",    foreground="#aaaaaa")
        # Order-group colours — cycle through these for distinct order grouping
        ORDER_PALETTE = [
            "#c9e8f9",  # 0 blue
            "#c9f0d4",  # 1 green
            "#fdf3c0",  # 2 yellow
            "#e8d8f5",  # 3 lavender
            "#fad9b0",  # 4 peach
            "#c5ede4",  # 5 teal
        ]
        for idx, color in enumerate(ORDER_PALETTE):
            self.tree.tag_configure(f"order_{idx}", background=color, foreground="#000000")

        style = ttk.Style()
        style.configure("Treeview",         font=("Helvetica", 12), rowheight=32,
                        foreground="#000000")
        style.configure("Treeview.Heading", font=("Helvetica", 12, "bold"))

        self.tree.bind("<Button-1>", self._on_click)

    # ── Open CSV ────────────────────────────────────────────────────────────────
    def _open_csv(self):
        path = filedialog.askopenfilename(
            title="Chọn file CSV TikTok Shop",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        if not path:
            return
        try:
            with open(path, newline="", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                raw = [{k.strip(): v.strip() for k, v in row.items()} for row in reader]
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không đọc được file:\n{e}")
            return

        self.items = filter_rows(raw, self.variant_map, self.size_fix, self.color_fix)
        if not self.items:
            messagebox.showinfo("Không có dữ liệu", "Không tìm thấy đơn 'To ship' nào.")
            return
        self._populate_table()
        self.export_btn.config(state="normal")

    # ── Populate ────────────────────────────────────────────────────────────────
    def _parse_order_date(self, date_str):
        """Try to parse a date string into a date object. Returns None on failure."""
        from datetime import datetime
        for fmt in (
            "%m/%d/%Y %I:%M:%S %p",   # 05/20/2026 7:43:26 PM  ← TikTok format
            "%m/%d/%Y %H:%M:%S",
            "%m/%d/%Y",
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d",
            "%d/%m/%Y",
            "%Y/%m/%d",
        ):
            try:
                return datetime.strptime(date_str.strip(), fmt).date()
            except (ValueError, AttributeError):
                continue
        return None

    def _populate_table(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        self.check_vars.clear()

        # Assign a palette index to each unique order_id in the order they appear
        order_color_idx = {}
        palette_size    = 6  # must match ORDER_PALETTE length in _build_ui

        for i, item in enumerate(self.items):
            oid = item["order_id"]
            is_locked   = not item["variant_id"]
            is_partial  = item.get("partial_lock", False)
            any_lock    = is_locked or is_partial
            var = tk.BooleanVar(value=not any_lock)
            self.check_vars.append(var)

            if oid not in order_color_idx:
                order_color_idx[oid] = len(order_color_idx) % palette_size

            if is_locked:
                tag = "locked"
            elif is_partial:
                tag = "partial_locked"
            else:
                tag = f"order_{order_color_idx[oid]}"

            check_sym  = "—" if any_lock else "☑"
            vid        = item["variant_id"] if item["variant_id"] else "—"
            lbl        = item["link_label"]   or PLACEHOLDER
            front      = item["design_front"] or PLACEHOLDER
            back       = item["design_back"]  or PLACEHOLDER
            mfront     = item["mockup_front"] or PLACEHOLDER
            mback      = item["mockup_back"]  or PLACEHOLDER

            self.tree.insert("", "end", iid=str(i), tags=(tag,), values=(
                check_sym, item["order_date"], item["order_id"], item["customer"],
                item["variation"], vid, item["quantity"],
                lbl, front, back, mfront, mback, item["status_note"],
            ))

        self._update_status(sum(1 for v in self.check_vars if v.get()))

    # ── Click ────────────────────────────────────────────────────────────────────
    def _on_click(self, event):
        if self._edit_entry:
            self._commit_edit()

        col = self.tree.identify_column(event.x)
        iid = self.tree.identify_row(event.y)
        if not iid:
            return

        col_names = ("check", "order_date", "order_id", "customer",
                     "variation", "variant_id", "qty",
                     "link_label", "design_front", "design_back",
                     "mockup_front", "mockup_back", "note")
        col_index = int(col.replace("#", "")) - 1
        col_name  = col_names[col_index] if col_index < len(col_names) else ""
        idx       = int(iid)

        # Checkbox toggle — block locked and partial_locked rows
        if col_name == "check":
            if not self.items[idx]["variant_id"] or self.items[idx].get("partial_lock"):
                return  # locked
            new_val = not self.check_vars[idx].get()
            self.check_vars[idx].set(new_val)
            vals    = list(self.tree.item(iid, "values"))
            vals[0] = "☑" if new_val else "☐"
            self.tree.item(iid, values=vals)
            self._update_status(sum(1 for v in self.check_vars if v.get()))
            return

        if col_name not in ("link_label", "design_front", "design_back",
                            "mockup_front", "mockup_back"):
            return

        try:
            bbox = self.tree.bbox(iid, col)
        except Exception:
            return
        if not bbox:
            return

        x, y, w, h = bbox
        current_val = self.items[idx][col_name]   # real value, not placeholder

        entry = tk.Entry(self.tree, font=("Helvetica", 12), relief="solid", bd=1,
                         highlightthickness=1, highlightbackground="#1565c0")
        entry.place(x=x, y=y, width=max(w, 300), height=h)
        entry.insert(0, current_val)
        entry.select_range(0, "end")
        entry.focus_set()

        self._edit_entry     = entry
        self._edit_iid       = iid
        self._edit_col_index = col_index
        self._edit_col_name  = col_name

        entry.bind("<Return>",    lambda e: self._commit_edit())
        entry.bind("<Escape>",    lambda e: self._cancel_edit())
        entry.bind("<FocusOut>",  lambda e: self._commit_edit())
        entry.bind("<KeyRelease>", lambda e: self._live_validate_entry())

    # ── Edit helpers ─────────────────────────────────────────────────────────────
    def _live_validate_entry(self):
        """Provide instant visual feedback while the user is typing a URL."""
        if not self._edit_entry:
            return
        val = self._edit_entry.get().strip()
        if is_valid_url(val):
            # URL looks good — restore normal appearance and hide the warning
            self._edit_entry.config(bg="white", highlightbackground="#1565c0")
            if hasattr(self, "_url_warn_label"):
                self._url_warn_label.place_forget()
        else:
            # Still invalid — keep red border
            self._edit_entry.config(bg="#ffd0d0", highlightbackground="#c0392b")

    def _commit_edit(self):
        if not self._edit_entry:
            return
        val        = self._edit_entry.get().strip()
        iid        = self._edit_iid
        idx        = int(iid)
        col_name   = self._edit_col_name
        col_index  = self._edit_col_index

        # URL validation
        if not is_valid_url(val):
            self._edit_entry.config(bg="#ffd0d0", highlightbackground="#c0392b")
            self._edit_entry.focus_set()
            if not hasattr(self, "_url_warn_label"):
                self._url_warn_label = tk.Label(
                    self.tree,
                    text="⚠ Link không hợp lệ (phải bắt đầu bằng https://...)",
                    bg="#ffd0d0", fg="#c0392b",
                    font=("Helvetica", 10), relief="solid", bd=1)
            bx, by, bw, bh = self.tree.bbox(iid, f"#{col_index + 1}")
            self._url_warn_label.place(x=bx, y=by + bh, width=340)
            self._url_warn_label.lift()
            return

        if hasattr(self, "_url_warn_label"):
            self._url_warn_label.place_forget()

        self.items[idx][col_name] = val

        # Recompute tag
        has_url_err = any(
            not is_valid_url(self.items[idx][f])
            for f in ("link_label", "design_front", "design_back",
                      "mockup_front", "mockup_back")
        )
        # Determine which order-group colour tag this row should have
        all_oids = [it["order_id"] for it in self.items]
        seen = {}
        for oid in all_oids:
            if oid not in seen:
                seen[oid] = len(seen) % 6
        order_tag = f"order_{seen[self.items[idx]['order_id']]}"

        current_tags = [t for t in self.tree.item(iid, "tags")
                        if t not in ("locked", "url_err") and not t.startswith("order_")]
        if has_url_err:
            current_tags.append("url_err")
        elif not self.items[idx]["variant_id"]:
            current_tags.append("locked")
        else:
            current_tags.append(order_tag)
        self.tree.item(iid, tags=current_tags)

        # Display: placeholder if empty
        display_val = val if val else PLACEHOLDER
        vals = list(self.tree.item(iid, "values"))
        vals[col_index] = display_val
        self.tree.item(iid, values=vals)

        self._edit_entry.destroy()
        self._edit_entry = None

    def _cancel_edit(self):
        if self._edit_entry:
            self._edit_entry.destroy()
            self._edit_entry = None

    # ── Select all / none ────────────────────────────────────────────────────────
    def _select_all(self):
        for i, var in enumerate(self.check_vars):
            if not self.items[i]["variant_id"] or self.items[i].get("partial_lock"):
                continue
            var.set(True)
            vals    = list(self.tree.item(str(i), "values"))
            vals[0] = "☑"
            self.tree.item(str(i), values=vals)
        self._update_status(sum(1 for v in self.check_vars if v.get()))

    def _select_none(self):
        for i, var in enumerate(self.check_vars):
            if not self.items[i]["variant_id"] or self.items[i].get("partial_lock"):
                continue
            var.set(False)
            vals    = list(self.tree.item(str(i), "values"))
            vals[0] = "☐"
            self.tree.item(str(i), values=vals)
        self._update_status(0)

    def _update_status(self, checked):
        total         = len(self.items)
        no_variant    = sum(1 for it in self.items if not it["variant_id"])
        partial       = sum(1 for it in self.items if it.get("partial_lock"))
        total_locked  = no_variant + partial
        self.status_var.set(
            f"Đã chọn {checked} / {total - total_locked} mặt hàng   |   "
            f"� {no_variant} thiếu Variant ID   🟠 {partial} đơn chưa đầy đủ"
        )

    # ── Export ───────────────────────────────────────────────────────────────────
    def _export(self):
        if self._edit_entry:
            self._commit_edit()

        # Guard: no partial orders — if an order has multiple items, either ALL
        # exportable items must be checked, or none of them.
        checked_indices = [i for i, v in enumerate(self.check_vars) if v.get()]
        partial_orders  = get_partial_export_violations(self.items, checked_indices)

        if partial_orders:
            messagebox.showerror(
                "Không thể xuất đơn hàng không đầy đủ",
                "❌ Các đơn hàng sau bị chọn thiếu sản phẩm.\n"
                "Vui lòng chọn tất cả sản phẩm trong đơn hoặc bỏ chọn toàn bộ:\n\n"
                + "\n".join(partial_orders)
            )
            return

        # Soft warning: invalid URLs in any link field (does NOT block export)
        invalid_urls = []
        for i, item in enumerate(self.items):
            if not self.check_vars[i].get():
                continue
            customer = item["customer"] or item["order_id"]
            for field, label in (
                ("link_label",   "Link Label"),
                ("design_front", "Mặt trước (Design Front)"),
                ("design_back",  "Mặt sau (Design Back)"),
                ("mockup_front", "Mockup Mặt trước"),
                ("mockup_back",  "Mockup Mặt sau"),
            ):
                val = item[field]
                if val and not is_valid_url(val):
                    invalid_urls.append(f"  • {customer} — {label}: \"{val}\"")

        if invalid_urls:
            proceed = messagebox.askyesno(
                "Link không hợp lệ",
                "⚠ Các link sau không hợp lệ (không bắt đầu bằng https://).\n"
                "Bạn có muốn tiếp tục xuất file không?\n\n"
                + "\n".join(invalid_urls)
            )
            if not proceed:
                return

        out_name = f"flashship_import_{date.today().isoformat()}.xlsx"
        out_path = os.path.join(SCRIPT_DIR, out_name)

        exported = skipped_unchecked = skipped_no_variant = 0

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "FlashShip Import"
        ws.append(FLASHSHIP_COLUMNS)

        from openpyxl.styles import Font, PatternFill, Alignment
        hdr_fill = PatternFill("solid", fgColor="1a1a2e")
        hdr_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for i, item in enumerate(self.items):
            if not self.check_vars[i].get():
                skipped_unchecked += 1
                continue
            if not item["variant_id"] or item.get("partial_lock"):
                skipped_no_variant += 1
                continue

            row = build_flashship_row(item)
            ws.append([row[col] for col in FLASHSHIP_COLUMNS])
            exported += 1

        # Auto-size columns
        for col_cells in ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0) for c in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 60)

        if exported == 0:
            messagebox.showwarning("Không có dữ liệu",
                                   "Không có hàng nào được xuất.\n"
                                   "Kiểm tra lại các hàng đã chọn có Variant ID hợp lệ.")
            return

        wb.save(out_path)
        msg = (f"✅  Đã xuất {exported} mặt hàng ra file:\n{out_path}\n\n"
               f"Bỏ qua — không chọn: {skipped_unchecked}\n"
               f"Bỏ qua — thiếu Variant ID: {skipped_no_variant}")
        messagebox.showinfo("Xuất file thành công", msg)
        os.system(f'open -R "{out_path}"')


if __name__ == "__main__":
    App().mainloop()
